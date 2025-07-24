# src/writers/excel_writer.py

import os, re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from ..core.logger import get_logger
from ..core.config import config

_INVALID_XML_CHARS_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

def _sanitize_for_excel(value):
    if isinstance(value, str):
        return _INVALID_XML_CHARS_RE.sub("", value)
    return value

class ExcelWriter:
    def __init__(self):
        self.log = get_logger(__name__)
        self.cfg = config._config["excel"]
        self.paths = config.paths

    def write_dataframe(self, df, out_path):
        out_path = self._prepare_path(out_path)
        wb = Workbook()
        ws = wb.active
        ws.title = self.cfg["default_sheet_name"]

        start = 1
        if self.cfg["include_header"]:
            for i, col in enumerate(df.columns, 1):
                ws.cell(1, i, _sanitize_for_excel(str(col)))
            start = 2

        for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), start):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, _sanitize_for_excel(val if pd.notna(val) else ""))

        if self.cfg["auto_adjust_width"]:
            for i, col in enumerate(df.columns, 1):
                length = max(len(str(col)), *(len(str(v)) for v in df[col])) + 2
                ws.column_dimensions[ws.cell(1, i).column_letter].width = min(50, max(10, length))

        wb.save(out_path)
        self.log.info(f"Guardado {Path(out_path).name}")
        return out_path

    def write_using_template(self, df, template_path, out_path):
        out_path = self._prepare_path(out_path)
        wb = load_workbook(template_path)
        ws = wb["Ficha"]

        # Named ranges and mapping
        col_map = {
            "codigo": "A",
            "descripcion": "B",
            "um": "C",
            "cantidad": "D",
            "precio_unitario": "E",
            "importe": "F",
        }

        # Borrar datos antiguos en rng_detalle
        # Asumimos fila inicial 5
        start_row = 5
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.value = None

        # Escribir detalle
        for i, (_, row) in enumerate(df.iterrows(), start=start_row):
            for field, col in col_map.items():
                val = row.get(field, "")
                ws[f"{col}{i}"] = val

        last_row = start_row + len(df) - 1
        # Fórmulas de totales
        ws[f"F{last_row+1}"] = f"=SUM(F{start_row}:F{last_row})"
        # Impuestos y margen (suponiendo celdas fijas en plantilla)
        # Ejemplo: B{last_row+2} impuesto SS 1.5%, B{last_row+3} margen 30%
        ws[f"B{last_row+2}"] = f"=F{last_row+1}*0.015"
        ws[f"B{last_row+3}"] = f"=F{last_row+1}*0.30"

        wb.save(out_path)
        self.log.info(f"Guardado con plantilla {Path(out_path).name}")
        return out_path

    def write_multiple_dataframes(self, dfs, out_path):
        out_path = self._prepare_path(out_path)
        wb = Workbook(); wb.remove(wb.active)
        for name, df in dfs.items():
            ws = wb.create_sheet(title=_sanitize_for_excel(name)[:31])
            start = 1
            if self.cfg["include_header"]:
                for i, col in enumerate(df.columns, 1):
                    ws.cell(1, i, _sanitize_for_excel(str(col)))
                start = 2
            for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), start):
                for c, val in enumerate(row, 1):
                    ws.cell(r, c, _sanitize_for_excel(val if pd.notna(val) else ""))
            if self.cfg["auto_adjust_width"]:
                for i, col in enumerate(df.columns, 1):
                    length = max(len(str(col)), *(len(str(v)) for v in df[col])) + 2
                    ws.column_dimensions[ws.cell(1, i).column_letter].width = min(50, max(10, length))
        wb.save(out_path)
        self.log.info(f"Guardado multi-hoja {Path(out_path).name}")
        return out_path

    def _prepare_path(self, path):
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        if not os.path.isabs(path):
            path = os.path.join(self.paths.output_dir, path)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        return path

